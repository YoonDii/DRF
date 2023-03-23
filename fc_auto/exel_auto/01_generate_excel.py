# 라이브러리 불러오기
from openpyxl import Workbook

# 클래스만들고 함수정리하기
class WeeklyWorkPlan:
    wb = None  # 워크북
    ws = None  # 워크시트
    start_date = "2023-03-01"
    manager = "매니저 이름을 입력 해주세요."

    def __init__(self, start_date, manager, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]  # 어떤시트를 열것인지
        self.start_date = start_date
        self.manager = manager
        self.set_title()

    # 파일 저장하는 함수
    def save(self, filename):
        self.wb.save(filename)
        print("엑셀파일 생성 완료")

    # 워크시트 셀선택과 입력값 설정
    def set_title(self):
        self.ws["B2"] = "담당자"
        self.ws["C2"] = self.manager
        self.ws["B3"] = "시작일"
        self.ws["C3"] = self.start_date


if __name__ == "__main__":  # 생성자 입력하기
    wwp = WeeklyWorkPlan("2023-03-23", "yoondii")
    wwp.save("주간업무계획표.xlsx")
